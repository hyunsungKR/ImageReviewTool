import sys
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QPushButton,
                               QLabel, QScrollArea, QTableWidget, QTableWidgetItem,
                               QHeaderView, QSizePolicy, QGridLayout, QSplitter)
from PySide6.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from openpyxl import load_workbook

EXCEL_FILE = "graphViewer/text.xlsx"

# CSV 파일 경로 지정 : 주석을 풀어서 사용하세요.

# CSV 파일 경로 : 1차 테스트
CSV_FILES = [
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC1_size1280_e200_b16_n167\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC2_size1280_e200_b16_n16\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC3_size1280_e200_b16_n16\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC4_size1280_e200_b16_n162\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC5_size1280_e200_b16_n16\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC6_size1280_e200_b16_n164\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC7_HY_size1280_e200_b16_n162\results.csv",
    # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC8_size1280_e300_b16_n16\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC9_size1280_e200_b16_n162\results.csv",
    r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC10_size1280_e200_b8_n16\results.csv"
]

# # CSV 파일 경로 : 1차 테스트 (일부)
# CSV_FILES = [
#     # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC1_size1280_e200_b16_n167\results.csv",
#     # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC2_size1280_e200_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC3_size1280_e200_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC4_size1280_e200_b16_n162\results.csv",
#     # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC5_size1280_e200_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC6_size1280_e200_b16_n164\results.csv",
#     # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC7_HY_size1280_e200_b16_n162\results.csv",
#     # # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC8_size1280_e300_b16_n16\results.csv",
#     # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC9_size1280_e200_b16_n162\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC10_size1280_e200_b8_n16\results.csv"
# ]

# # CSV 파일 경로 : 2차 테스트
# CSV_FILES = [
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC1_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC2_training2\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC3_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC5_training2\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC6_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC7_training3\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC8_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC10_training4\results.csv"
# ]

# # CSV 파일 경로 : 3차 테스트
# CSV_FILES = [
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC1_training4\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC2_training3\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC3_training3\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC5_training5\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC6_training\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC8_training2\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC10_training\results.csv",
# ]

# # CSV 파일 경로 : 4차 테스트
# CSV_FILES = [
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC2_1_training\results.csv",
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC3_1_training\results.csv",
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC6_1_training\results.csv",
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC8_1_training\results.csv"
# ]

# # CSV 파일 경로 : 모든 결과값
# CSV_FILES = [
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC1_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC2_training2\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC3_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC5_training2\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC6_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC7_training3\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC8_training\results.csv",
#     r"\\tnsai-khy\test\240926TEST_RESULT\PC10_training4\results.csv",
#
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC1_size1280_e200_b16_n167\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC2_size1280_e200_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC3_size1280_e200_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC4_size1280_e200_b16_n162\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC5_size1280_e200_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC6_size1280_e200_b16_n164\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC7_HY_size1280_e200_b16_n162\results.csv",
#     # r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC8_size1280_e300_b16_n16\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC9_size1280_e200_b16_n162\results.csv",
#     r"\\tnsai-khy\test\240925TEST_RESULT\240925_swhTest_PC10_size1280_e200_b8_n16\results.csv",
#
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC1_training4\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC2_training3\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC3_training3\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC5_training5\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC6_training\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC8_training2\results.csv",
#     r"\\tnsai-khy\test\240927TEST_RESULT\PC10_training\results.csv",
#
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC2_1_training\results.csv",
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC3_1_training\results.csv",
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC6_1_training\results.csv",
#     r"\\tnsai-khy\test\240929TEST_RESULT\PC8_1_training\results.csv"
# ]

# 컬럼 설명
COLUMN_DESCRIPTIONS = {
    "train/box_loss": "(Bounding box loss: 정확한 객체 위치 예측)",
    "train/cls_loss": "(Class loss: 클래스 분류 정확도)",
    "train/dfl_loss": "(Distribution focal loss: 확률 분포 손실)",
    "metrics/precision": "(Precision: 예측된 객체의 정밀도)",
    "metrics/recall": "(Recall: 모델이 감지한 객체의 비율)",
    "metrics/mAP50": "(mAP50: IoU 0.50에서의 모델 성능)",
    "metrics/mAP50-95": "(mAP50-95: IoU 0.50~0.95에서의 평균 정밀도)",
    "val/box_loss": "(Validation box loss: 검증 데이터셋의 바운딩 박스 손실)",
    "val/cls_loss": "(Validation class loss: 검증 데이터셋의 클래스 손실)",
    "val/dfl_loss": "(Validation distribution focal loss: 검증 데이터셋의 분포 손실)",
    "lr/pg0": "(Learning rate pg0: 첫 번째 매개변수 그룹의 학습률)",
    "lr/pg1": "(Learning rate pg1: 두 번째 매개변수 그룹의 학습률)",
    "lr/pg2": "(Learning rate pg2: 세 번째 매개변수 그룹의 학습률)"
}


class GraphViewer(QWidget):
    def __init__(self):
        super().__init__()
        self.layout = QVBoxLayout(self)

        # 설명글 섹션
        self.description_label = QLabel("설명 파일을 로드하세요.", self)
        self.description_label.setAlignment(Qt.AlignTop)
        self.description_label.setWordWrap(True)
        self.description_label.setStyleSheet("font-size: 15px;")
        self.description_label.setTextFormat(Qt.RichText)  # HTML 형식 사용

        # 설명 스크롤
        self.scroll = QScrollArea(self)
        self.scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.addWidget(self.description_label)
        self.scroll.setWidget(scroll_content)
        self.scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        # 그래프 캔버스
        self.canvas = FigureCanvas(plt.Figure(figsize=(10, 6)))
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.ax = self.canvas.figure.add_subplot(111)

        # 통계 테이블
        self.stats_table = QTableWidget(0, 3, self)
        self.stats_table.setHorizontalHeaderLabels(['File', 'Max', 'Min'])
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        self.stats_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.stats_table.setFont(QFont("Arial", 15))
        self.stats_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 그래프와 테이블, 설명을 위한 스플리터
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self.scroll)
        splitter.addWidget(self.canvas)
        splitter.addWidget(self.stats_table)
        splitter.setSizes([150, 400, 200])  # 초기 크기 조정
        self.layout.addWidget(splitter)

        # 버튼 레이아웃
        self.button_layout = QGridLayout()
        self.button_layout.setColumnStretch(0, 1)
        self.button_layout.setColumnStretch(1, 1)

        # 버튼 레이아웃을 담을 위젯
        button_widget = QWidget()
        button_widget.setLayout(self.button_layout)
        button_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.layout.addWidget(button_widget)

        # 판다스 출력 옵션 설정: 부동소수점 표시 자릿수 늘리기
        pd.set_option('display.float_format', lambda x: '%.15g' % x)

        self.load_description(EXCEL_FILE)
        self.create_buttons()



    def load_description(self, excel_path):
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        description_html = ""

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            row_html = ""
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value)
                    cell_font = cell.font

                    # HTML 태그로 폰트 스타일 적용
                    styles = []
                    if cell_font.bold:
                        styles.append('font-weight:bold;')
                    if cell_font.italic:
                        styles.append('font-style:italic;')
                    # 색상 처리
                    if cell_font.color:
                        color = cell_font.color
                        if color.type == 'rgb' and color.rgb:
                            color_hex = color.rgb
                            if len(color_hex) == 8:  # ARGB 형식인 경우 알파 채널 제거
                                color_hex = color_hex[2:]
                            styles.append(f'color:#{color_hex};')
                        elif color.type == 'theme':
                            # 테마 색상 처리 (필요한 경우)
                            pass
                        elif color.type == 'indexed':
                            # 인덱스 색상 처리 (필요한 경우)
                            pass
                    if cell_font.size:
                        styles.append(f'font-size:{cell_font.size}pt;')

                    style_str = ''.join(styles)
                    cell_html = f'<span style="{style_str}">{cell_text}</span>'

                    row_html += cell_html + ' '
            description_html += row_html + '<br>'

        self.description_label.setText(description_html)

    def create_buttons(self):
        """각 컬럼에 대해 버튼 생성"""
        row = 0
        col = 0
        for column, description in COLUMN_DESCRIPTIONS.items():
            button = QPushButton(f"{column} : {description}")
            button.setStyleSheet("text-align:left; padding: 12px; font-size: 15px; font-weight: bold;")
            button.clicked.connect(lambda checked, col=column: self.plot_graph(col))
            self.button_layout.addWidget(button, row, col)
            col += 1
            if col == 2:
                col = 0
                row += 1

    def plot_graph(self, column_name):
        """그래프를 그리는 함수"""
        self.ax.clear()
        colors = plt.cm.viridis(np.linspace(0, 1, len(CSV_FILES)))
        self.stats_table.setRowCount(len(CSV_FILES))  # 테이블 행 수 재설정
        plotted = False

        for i, file_path in enumerate(CSV_FILES):
            dir_name = os.path.basename(os.path.dirname(file_path))  # 디렉토리 이름 추출
            try:
                # CSV 파일 읽기
                df = pd.read_csv(file_path)
            except Exception as e:
                print(f"{dir_name}의 파일을 읽는 중 오류 발생: {e}")
                continue

            # 컬럼 이름을 소문자로 변환하고 공백 제거
            df.columns = [col.lower().replace(' ', '') for col in df.columns]
            # 목표 컬럼 이름도 동일하게 처리
            target_col = column_name.lower().replace(' ', '')

            # 부분적으로 일치하는 컬럼 찾기
            matched_columns = [col for col in df.columns if target_col in col]
            if not matched_columns:
                print(f"{dir_name}에서 {column_name} 컬럼을 찾을 수 없습니다.")
                continue

            # 첫 번째 컬럼은 에포크, 그 이후가 데이터
            epoch_data = df.iloc[:, 0]
            data = df[matched_columns[0]]

            # 데이터 숫자로 변환
            try:
                epoch_data = pd.to_numeric(epoch_data, errors='coerce')
                data = pd.to_numeric(data, errors='coerce')
            except Exception as e:
                print(f"{dir_name}의 데이터를 숫자로 변환하는 중 오류 발생: {e}")
                continue

            # NaN 값 제거
            valid_indices = data.notna() & epoch_data.notna()
            epoch_data = epoch_data[valid_indices]
            data = data[valid_indices]

            if data.empty or epoch_data.empty:
                print(f"{dir_name}: {column_name} 데이터가 비어 있습니다.")
                continue

            # 데이터 타입 확인
            if not np.issubdtype(data.dtype, np.floating):
                print(f"{dir_name}: {column_name} 데이터가 부동소수점 타입이 아닙니다.")
                continue

            # 그래프 그리기
            self.ax.plot(epoch_data, data, label=dir_name, color=colors[i])
            max_val = data.max()
            min_val = data.min()
            max_idx = data.idxmax()
            min_idx = data.idxmin()

            # 그래프에 최대/최소 값 표시
            self.ax.plot(epoch_data.iloc[max_idx], max_val, 'o', color=colors[i])
            self.ax.plot(epoch_data.iloc[min_idx], min_val, 'o', color=colors[i])

            # 통계 테이블에 값 설정 (자릿수 제한 없이 표시)
            max_str = format(max_val, '.15g')
            min_str = format(min_val, '.15g')

            self.stats_table.setItem(i, 0, QTableWidgetItem(dir_name))  # 디렉토리 이름
            self.stats_table.setItem(i, 1, QTableWidgetItem(max_str))  # 최대값
            self.stats_table.setItem(i, 2, QTableWidgetItem(min_str))  # 최소값

            plotted = True

        if plotted:
            self.ax.set_title(f"{column_name} over epochs", fontsize=16)
            self.ax.set_xlabel("Epoch", fontsize=14)
            self.ax.set_ylabel("Value", fontsize=14)
            self.ax.ticklabel_format(style='plain', axis='y')  # Y축에 과학적 표기법 사용 안 함
            self.ax.legend(fontsize=12)
            self.ax.grid(True)
            self.ax.relim()              # 축 범위 재설정
            self.ax.autoscale()          # 축 스케일 자동 조정
            self.ax.margins(y=0.1)       # Y축에 여백 추가
        else:
            self.ax.set_title(f"No data for {column_name}", fontsize=16)
        self.canvas.draw()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = GraphViewer()
    window.setWindowTitle("YOLOv8 Training Results Viewer")
    window.resize(1200, 800)
    window.showMaximized()
    window.show()
    sys.exit(app.exec())
